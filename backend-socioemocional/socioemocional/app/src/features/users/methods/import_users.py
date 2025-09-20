from collections import Counter
from typing import List
from fastapi import HTTPException, Depends, UploadFile
from sqlalchemy.orm import Session
from src.domains.team import Team
from src.domains.user import User
from src.domains.enums.role_type import RoleType
from src.infrastructure.results.import_users import ImportUserResult, ImportUserErrorResult
from src.infrastructure.validations.fields import is_valid_email, is_cpf
from src.infrastructure.utils import get_column_values_by_header, remove_special_characters
from src.data.database import get_db
from pydantic import BaseModel
from src.domains.enums.import_users_cells import ImportUsersCells
from openpyxl import load_workbook
import numpy as np
from . import BaseHandler

# Request
class Command(BaseModel):
    pass

class UsersToAdd:
    def __init__(self, name, email, cpf, role, team_string, manager_team):
        self.name: str = name
        self.email: str = email
        self.cpf: str = cpf
        self.role: int = role
        self.team_string: str = team_string
        self.manager_team: str = manager_team

# Handle
class ImportUsers(BaseHandler[Command, ImportUserResult]):
    def __init__(self, db: Session = Depends(get_db)):
        self.db = db

    async def execute(self, file: UploadFile):
        if not file.filename.endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="O arquivo deve ser .xlsx")
        
        wb = load_workbook(filename=file.file)
        ws = wb.active

        cpfs = [remove_special_characters(str(cpf)) for cpf in get_column_values_by_header(ws, ImportUsersCells.CPF.value)]
        cpf_counts = Counter(cpfs)
        duplicated_cpfs = [cpf for cpf, count in cpf_counts.items() if count > 1]

        existing_users = (self.db.query(User)
                 .filter(User.cpf.in_(np.unique(cpfs)))
                 .all())

        users_errors: List[ImportUserErrorResult] = []
        users_to_add: List[UsersToAdd] = []

        # Validação
        for row in ws.iter_rows(min_row=2, values_only=True): # Pula cabeçalho
            name, email, cpf, role, team_string, manager_team = row

            if cpf:
                cpf = remove_special_characters(str(cpf))

            if not name or not cpf or not role or not team_string or not manager_team:
                users_errors.append(ImportUserErrorResult(nome=name, email=email, cpf=cpf, cargo=role, equipe=team_string, gerenteEquipe=manager_team, errors=["Alguns dos campos obrigatórios estão vazios."]))
                continue

            if cpf in duplicated_cpfs:
                users_errors.append(ImportUserErrorResult(nome=name, email=email, cpf=cpf, cargo=role, equipe=team_string, gerenteEquipe=manager_team, errors=["CPF duplicado."]))
                continue
            
            # Semelhante ao FirstOrDefault do C#, busca se um usuário existe por cpf, do contrário retorna None
            if next((u for u in existing_users if u.cpf == cpf), None):
                users_errors.append(ImportUserErrorResult(nome=name, email=email, cpf=cpf, cargo=role, equipe=team_string, gerenteEquipe=manager_team, errors=[f"Usuário com CPF {cpf} já existe no banco de dados."]))
                continue
            
            errors = self.validate_fields(email, cpf, role)
            if len(errors) > 0:
                users_errors.append(ImportUserErrorResult(nome=name, email=email, cpf=cpf, cargo=role, equipe=team_string, gerenteEquipe=manager_team, errors=errors))
                continue

            users_to_add.append(UsersToAdd(name, email, cpf, role, team_string, manager_team))
        
        team_names = np.unique([team_name.lower() for team_name in get_column_values_by_header(ws, ImportUsersCells.Equipe.value)])
        existing_teams = (self.db.query(Team)
                 .filter(Team.name.in_(team_names))
                 .all())
        
        existing_teams_dict = {t.name.lower(): t for t in existing_teams}

        teams_to_add: List[Team] = []
        if len(users_to_add) > 0:
            for team_name in team_names:
                users = [user for user in users_to_add if user.team_string.lower() == team_name and user.manager_team.lower() != "s"]
                managers = [user for user in users_to_add if user.team_string.lower() == team_name and user.manager_team.lower() == "s"]

                users = [User(u.name, u.email, u.cpf, RoleType(u.role)) for u in users]
                managers = [User(u.name, u.email, u.cpf, RoleType(u.role)) for u in managers]

                if len(users) == 0 and len(managers) == 0:
                    continue

                if team_name in existing_teams_dict:
                    team = existing_teams_dict[team_name]
                    team.users.extend(users)
                    team.managers.extend(managers)
                else:
                    team = Team(team_name)
                    team.users = users
                    team.managers = managers
                    teams_to_add.append(team)

            self.db.add_all(teams_to_add)
            self.db.commit()

        return ImportUserResult(imported_users_count=len(users_to_add), error_import_users=users_errors)

    def validate_fields(self, email: str, cpf: str, role: int):
        errors = []
        columns = [cell.value for cell in RoleType]

        if not is_valid_email(email):
            errors.append("Email inválido.")

        if not is_cpf(cpf):
            errors.append("CPF inválido.")

        if role not in columns:
            errors.append("Cargo inválido.")
        
        return errors